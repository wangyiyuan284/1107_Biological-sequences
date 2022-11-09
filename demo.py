# -*- coding: utf-8 -*-
import requests
import re
import openpyxl
from bs4 import BeautifulSoup

# Get user's inputet

proteinFamily = str(input("Please enter the specified protein family:"))
taxonomicGroup = str(input("Please enter the specified taxonomic group:"))

head = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36'}
url = r'https://www.ncbi.nlm.nih.gov/protein/?term='
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'],ws['B1'],ws['C1'],ws['D1'],ws['E1']  = '假定蛋白信息','得分','结构域1','结构域2','结构域3'
ws.freeze_panes = 'A2'

def seq_data(file, filename):
    for line in open(file):
        if 'MGG_' in line:
            score = line.split(',')[2]
            MGG = re.search(r'MGG_\d{5}',line).group(0)
            full_url = url + MGG
            l = []
            l.append(MGG)
            l.append(score)
            try:
                res = requests.get(full_url,headers = head)
                res.encoding = 'utf-8'
                soup = BeautifulSoup(res.text,'lxml')
                domain = soup.find_all("dd",class_='clearfix')#获取标签内容
                for each in domain:
                    l.append(each.text)
            except BaseException:
                pass
            ws.append(l)#写入excel
    wb.save(filename)#保存
            
                     
if __name__ == '__main__':            
    seq_data(r'C:\Users\zhuxueming\Desktop\unique_Atg31.csv',\
         r'C:\Users\zhuxueming\Desktop\ATG3_special_hyp_protein_domain.xlsx')